# ============================================================
# backend/server/api/routes/report.py
# Report 保存/列举/加载/导出
# ============================================================

import json
from collections import Counter
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query

from server.api.schemas.report import ReportSaveRequest, ReportListItem

router = APIRouter(prefix="/api/report", tags=["report"])

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
REPORT_DIR = PROJECT_ROOT / "Report_Data"


@router.post("/save", status_code=200)
def save_report(body: ReportSaveRequest):
    report_dir = REPORT_DIR / body.filename.replace('.json', '')
    report_dir.mkdir(parents=True, exist_ok=True)
    (report_dir / 'images').mkdir(exist_ok=True)

    data = body.data
    for det in data.get('defects', []):
        ai = det.pop('annotated_image', None) or det.pop('annotatedImage', None)
        if ai:
            det['image'] = ai

    (report_dir / 'metadata.json').write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    return {"status": "ok", "dir": report_dir.name}


@router.get("/list", response_model=list[ReportListItem])
def list_reports():
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    items: list[ReportListItem] = []
    for entry in sorted(REPORT_DIR.iterdir(), reverse=True):
        if entry.is_dir() and (entry / 'metadata.json').exists():
            items.append(ReportListItem(filename=entry.name))
    return items


@router.get("/{name:path}")
def load_report(name: str):
    p = REPORT_DIR / name / 'metadata.json'
    if not p.exists():
        raise HTTPException(404, f"Report {name} 不存在")
    return json.loads(p.read_text(encoding='utf-8'))


# ============================================================
# 导出
# ============================================================

@router.post("/{name:path}/export")
def export_report(name: str, format: str = Query("md")):
    p = REPORT_DIR / name / 'metadata.json'
    if not p.exists():
        raise HTTPException(404, f"Report {name} 不存在")
    data = json.loads(p.read_text(encoding='utf-8'))
    report_dir = REPORT_DIR / name
    defects = data.get('defects', [])
    task = data.get('task_name', '')
    date = data.get('date', '')
    pcurl = data.get('point_cloud_url', '')

    if format == 'md':
        return _gen_markdown(report_dir, defects, task, date, pcurl)
    elif format == 'xlsx':
        return _gen_xlsx(report_dir, defects, task, date, pcurl)
    else:
        raise HTTPException(400, f"不支持的格式: {format}")


def _gen_markdown(report_dir: Path, defects: list, task: str, date: str, pcurl: str):
    # 按类型统计
    type_counts = Counter(d.get('class_name', '未知') for d in defects)

    lines = [
        f"# 缺陷检测报告 — {task} ({date})",
        "",
        "## 概要",
        f"- 检测任务: {task}",
        f"- 检测日期: {date}",
        f"- 缺陷总数: {len(defects)} 条",
    ]
    if pcurl:
        lines.append(f"- 点云文件: {pcurl}")

    # 分类统计
    lines.append("")
    lines.append("### 分类统计")
    lines.append("| 类型 | 数量 |")
    lines.append("|------|------|")
    for t, c in sorted(type_counts.items(), key=lambda x: -x[1]):
        lines.append(f"| {t} | {c} |")

    # 缺陷表
    lines.append("")
    lines.append("## 缺陷列表")
    lines.append("| ID | 类型 | 置信度 | 位置 (X, Y, Z) | 标注图 |")
    lines.append("|----|------|--------|----------------|--------|")
    for d in defects:
        did = d.get('id', '-')
        cls = d.get('class_name', '-')
        conf = f"{int(d.get('confidence', 0) * 100)}%"
        c3 = d.get('center_3d', [0, 0, 0]) or [0, 0, 0]
        pos = f"{c3[0]:.2f}, {c3[1]:.2f}, {c3[2]:.2f}"
        img = d.get('image', '')
        img_cell = f"![图]({img})" if img else "-"
        lines.append(f"| {did} | {cls} | {conf} | {pos} | {img_cell} |")

    outpath = report_dir / 'report.md'
    outpath.write_text('\n'.join(lines), encoding='utf-8')
    return {"status": "ok", "path": "report.md"}


def _gen_xlsx(report_dir: Path, defects: list, task: str, date: str, pcurl: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet 1: 缺陷列表
    ws = wb.active
    ws.title = "缺陷列表"
    headers = ["序号", "ID", "类型", "置信度", "位置 X (m)", "位置 Y (m)", "位置 Z (m)", "标注图路径"]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    sorted_defects = sorted(defects, key=lambda d: d.get('confidence', 0), reverse=True)
    for i, d in enumerate(sorted_defects, 1):
        c3 = d.get('center_3d', [0, 0, 0]) or [0, 0, 0]
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=d.get('id', '-'))
        ws.cell(row=i + 1, column=3, value=d.get('class_name', '-'))
        ws.cell(row=i + 1, column=4, value=f"{int(d.get('confidence', 0) * 100)}%")
        ws.cell(row=i + 1, column=5, value=round(c3[0], 3))
        ws.cell(row=i + 1, column=6, value=round(c3[1], 3))
        ws.cell(row=i + 1, column=7, value=round(c3[2], 3))
        ws.cell(row=i + 1, column=8, value=d.get('image', ''))

    ws.freeze_panes = 'A2'
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Sheet 2: 概要
    ws2 = wb.create_sheet("概要")
    type_counts = Counter(d.get('class_name', '未知') for d in defects)
    info = [
        ("检测任务", task),
        ("检测日期", date),
        ("缺陷总数", len(defects)),
        ("点云文件", pcurl or "无"),
        ("", ""),
        ("类型", "数量"),
    ]
    for t, c in sorted(type_counts.items(), key=lambda x: -x[1]):
        info.append((t, c))

    for row, (k, v) in enumerate(info, 1):
        ws2.cell(row=row, column=1, value=str(k))
        ws2.cell(row=row, column=2, value=str(v))
        if row <= 4:
            ws2.cell(row=row, column=1).font = header_font

    outpath = report_dir / 'report.xlsx'
    wb.save(str(outpath))
    return {"status": "ok", "path": "report.xlsx"}
